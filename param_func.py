from openpyxl import load_workbook

def param_func(paramData):
	
	#this dictionary creates a key-value pair for each row of paramData (key to well)
	key = 0	
	paramRow = {}
	while key < len(paramData):
		paramRow[key] = paramData[key]
		key += 1

	#this dictionary creates a key-value pair for each parameter (key to parameter)
	param = {"wellID" : 0, "time" : 1, "rate" : 4, "DTW" : 5, "pH" : 7, "conductivity" : 8, "DO" : 11, "temp" : 12, "redox" : 13}

	index = 0
	nextIndex = index + 1
	pos = 0

	currentWell = paramRow[index][param["wellID"]]
	nextWell = paramRow[index + 1][param["wellID"]]

	#the following statements populate lowFlowForm template with param info
	for row in paramData:
		if index == 0:
			currentWell = paramRow[index][param["wellID"]]
			nextWell = paramRow[nextIndex][param["wellID"]]
			
			newbook = load_workbook(filename="GWS_form_" + currentWell + ".xlsx")
			newSheet = newbook.active

			#the following if/else statements prevent stray 0s from appearing
			if newSheet['C' + str(pos + 26)].value != "0":
				newSheet['A' + str(pos + 26)] = paramRow[index][param["time"]]

			else:
				newSheet['C' + str(pos + 26)].number_format = ''

			newSheet['E' + str(pos + 26)] = paramRow[index][param["rate"]]
			newSheet['I' + str(pos + 26)] = paramRow[index][param["temp"]]
			newSheet['K' + str(pos + 26)] = paramRow[index][param["pH"]]
			newSheet['M' + str(pos + 26)] = paramRow[index][param["conductivity"]]
			newSheet['O' + str(pos + 26)] = paramRow[index][param["redox"]]
			newSheet['Q' + str(pos + 26)] = paramRow[index][param["DO"]]
			newSheet['U' + str(pos + 26)] = paramRow[index][param["DTW"]]
			newbook.save(filename="GWS_form_" + currentWell + ".xlsx")

			index += 1
			nextIndex = index + 1
			pos += 1

		elif currentWell != nextWell:
			currentWell = paramRow[index][param["wellID"]]
			if nextIndex > 69:
				nextWell = currentWell

			else:
				nextWell = paramRow[nextIndex][param["wellID"]]
			pos = 0
			
			newbook = load_workbook(filename="GWS_form_" + currentWell + ".xlsx")
			newSheet = newbook.active
			
			#the following if/else statements prevent stray 0s from appearing
			if newSheet['C' + str(pos + 26)].value != "0":
				newSheet['A' + str(pos + 26)] = paramRow[index][param["time"]]

			else:
				newSheet['C' + str(pos + 26)].number_format = ''

			newSheet['E' + str(pos + 26)] = paramRow[index][param["rate"]]
			newSheet['I' + str(pos + 26)] = paramRow[index][param["temp"]]
			newSheet['K' + str(pos + 26)] = paramRow[index][param["pH"]]
			newSheet['M' + str(pos + 26)] = paramRow[index][param["conductivity"]]
			newSheet['O' + str(pos + 26)] = paramRow[index][param["redox"]]
			newSheet['Q' + str(pos + 26)] = paramRow[index][param["DO"]]
			newSheet['U' + str(pos + 26)] = paramRow[index][param["DTW"]]
			newbook.save(filename="GWS_form_" + currentWell + ".xlsx")

			index += 1
			nextIndex = index + 1
			pos += 1
		
		else:
			currentWell = paramRow[index][param["wellID"]]
			if nextIndex > 69:
				nextWell = currentWell

			else:
				nextWell = paramRow[nextIndex][param["wellID"]]
			
			newbook = load_workbook(filename="GWS_form_" + currentWell + ".xlsx")
			newSheet = newbook.active
			
			#the following if/else statements prevent stray 0s from appearing
			if newSheet['C' + str(pos + 26)].value != "0":
				newSheet['A' + str(pos + 26)] = paramRow[index][param["time"]]

			else:
				newSheet['C' + str(pos + 26)].number_format = ''
			
			newSheet['E' + str(pos + 26)] = paramRow[index][param["rate"]]
			newSheet['I' + str(pos + 26)] = paramRow[index][param["temp"]]
			newSheet['K' + str(pos + 26)] = paramRow[index][param["pH"]]
			newSheet['M' + str(pos + 26)] = paramRow[index][param["conductivity"]]
			newSheet['O' + str(pos + 26)] = paramRow[index][param["redox"]]
			newSheet['Q' + str(pos + 26)] = paramRow[index][param["DO"]]
			newSheet['U' + str(pos + 26)] = paramRow[index][param["DTW"]]
			newbook.save(filename="GWS_form_" + currentWell + ".xlsx")

			index += 1
			nextIndex = index + 1
			pos += 1