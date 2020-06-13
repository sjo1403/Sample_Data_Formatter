from openpyxl import load_workbook
from header_func import header_func
from param_func import param_func

sourcebook = load_workbook(filename="raw_data.xlsx")
sourcebook.active = 0
sheet1 = sourcebook.active	#workbook with raw data

#the next two statements populate the lowFlowForm with data found in the header
headerData = []

for row in sheet1.iter_rows (min_row=2, max_row=9, min_col=15, max_col=57, values_only=True):
	headerData.append(row)

header_func(headerData)

#the next two statements populate the lowFlowForm with data found in the field parameters section
paramData = []

sourcebook.active = 1
sheet2 = sourcebook.active

for row in sheet2.iter_rows (min_row=2, max_row=71, min_col=9, max_col=22, values_only=True):
	paramData.append(row)

param_func(paramData)
#constituent_func(stringData)
